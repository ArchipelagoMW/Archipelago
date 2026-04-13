#!/usr/bin/env python3
"""Iterative progression sphere analysis for CVAOS with map generation."""

from collections import deque
from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path
import sys


def _load_cvaos_data():
    # Load the standalone data package without importing worlds.cvaos,
    # which currently fails earlier during world registration.
    data_dir = Path(__file__).resolve().parent / "worlds" / "cvaos" / "data"
    spec = spec_from_file_location(
        "cvaos_data",
        data_dir / "__init__.py",
        submodule_search_locations=[str(data_dir)],
    )
    if spec is None or spec.loader is None:
        raise ImportError(f"Could not load CVAoS data package from {data_dir}")

    module = module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


_cvaos_data = _load_cvaos_data()
entrance_info_collection = _cvaos_data.entrance_info_collection
entrance_to_entrance_info_collection = _cvaos_data.entrance_to_entrance_info_collection
entrance_to_pickup_region_info_collection = _cvaos_data.entrance_to_pickup_region_info_collection
pickup_info_collection = _cvaos_data.pickup_info_collection
transdoor_connection_collection = _cvaos_data.transdoor_connection_collection
AbilityCombo = _cvaos_data.AbilityCombo

# Progression items (ability souls)
PROGRESSION_ITEMS = {
    'Flying Armor': AbilityCombo.Glide,
    'Skeleton Blaze': AbilityCombo.Slide,
    'Malphas': AbilityCombo.DJump,
    'Hippogryph': AbilityCombo.HJump,
    'Undine': AbilityCombo.WWalk,
    'Skula': AbilityCombo.Dive,
    'Black Panther': AbilityCombo.Panth,
    'Giant Bat': AbilityCombo.Bat,
    'Grave Keeper': AbilityCombo.BDash,
    'Kicker Skeleton': AbilityCombo.Kick,
}

# Build pickup_number -> item_name lookup
pickup_to_item = {p.pickup_number: p.simple_name for p in pickup_info_collection}

# Build door lookups
door_id_unique_to_number = {e.door_identifier_unique: e.door_number for e in entrance_info_collection}
door_number_to_room = {e.door_number: e.room_identifier for e in entrance_info_collection}

# Build entrance adjacency
entrance_adjacency: dict[int, list[tuple[int, tuple]]] = {}
for transdoor in transdoor_connection_collection:
    from_door_number = door_id_unique_to_number.get(transdoor.from_entrance)
    to_door_number = door_id_unique_to_number.get(transdoor.to_entrance)
    if from_door_number is not None and to_door_number is not None:
        entrance_adjacency.setdefault(from_door_number, []).append((to_door_number, (0,)))

for routing in entrance_to_entrance_info_collection:
    from_door_id = f'{routing.room_id}:{routing.from_room}'
    to_door_id = f'{routing.room_id}:{routing.to_room}'
    from_door_number = door_id_unique_to_number.get(from_door_id)
    to_door_number = door_id_unique_to_number.get(to_door_id)
    if from_door_number is not None and to_door_number is not None:
        masks = routing.get_requirement_bitmasks()
        entrance_adjacency.setdefault(from_door_number, []).append((to_door_number, masks))

# Build entrance -> pickups lookup
entrance_to_pickups: dict[int, list[tuple[int, tuple]]] = {}
for route in entrance_to_pickup_region_info_collection:
    door_number = door_id_unique_to_number.get(route.entrance_identifier)
    if door_number is not None:
        masks = route.get_requirement_bitmasks()
        entrance_to_pickups.setdefault(door_number, []).append((route.pickup_number, masks))


def can_traverse(masks, abilities_mask):
    if not masks:
        return True
    for mask in masks:
        if mask == 0:
            return True
        if mask & AbilityCombo.Impossible:
            continue
        if (mask & abilities_mask) == mask:
            return True
    return False


def get_reachable_data(abilities_mask):
    """BFS to find all reachable doors and pickups given abilities."""
    start_door_number = door_id_unique_to_number.get('000:003')
    if start_door_number is None:
        return set(), set()

    visited_doors = set([start_door_number])
    queue = deque([start_door_number])
    reachable_pickups = set()

    while queue:
        current = queue.popleft()

        # Check pickups reachable from this entrance
        for pickup_num, masks in entrance_to_pickups.get(current, []):
            if can_traverse(masks, abilities_mask):
                reachable_pickups.add(pickup_num)

        # Traverse to other entrances
        for dest, masks in entrance_adjacency.get(current, []):
            if dest not in visited_doors and can_traverse(masks, abilities_mask):
                visited_doors.add(dest)
                queue.append(dest)

    # Convert doors to rooms
    reachable_rooms = {door_number_to_room[d] for d in visited_doors if d in door_number_to_room}

    return reachable_rooms, reachable_pickups


def main():
    # Collect sphere data
    spheres = []
    abilities_mask = 0
    collected_items: set[str] = set()

    print('=== ITERATIVE PROGRESSION SPHERE ANALYSIS ===')
    print()

    while True:
        reachable_rooms, reachable_pickups = get_reachable_data(abilities_mask)

        # Find new progression items we can reach
        new_progression = []
        for pickup_num in reachable_pickups:
            item_name = pickup_to_item.get(pickup_num)
            if item_name in PROGRESSION_ITEMS and item_name not in collected_items:
                new_progression.append(item_name)

        abilities_str = ', '.join(sorted(collected_items)) if collected_items else '(none)'
        sphere_num = len(spheres)

        print(f'Sphere {sphere_num}:')
        print(f'  Abilities: {abilities_str}')
        print(f'  Reachable rooms: {len(reachable_rooms)}')
        print(f'  Reachable pickups: {len(reachable_pickups)}')
        print(f'  New progression items: {new_progression if new_progression else "(none)"}')
        print()

        spheres.append({
            'sphere': sphere_num,
            'abilities': abilities_str,
            'rooms': sorted(reachable_rooms),
            'new_items': new_progression,
        })

        if not new_progression:
            break

        for item in new_progression:
            collected_items.add(item)
            abilities_mask |= int(PROGRESSION_ITEMS[item])

    # Generate map tabs
    print('=== GENERATING MAP TABS ===')
    import subprocess
    import sys

    for sphere_data in spheres:
        sphere_num = sphere_data['sphere']
        rooms = sphere_data['rooms']
        sheet_name = f"Sphere {sphere_num}"
        rooms_str = ','.join(rooms)

        print(f'Creating tab: {sheet_name} ({len(rooms)} rooms)')

        result = subprocess.run(
            [sys.executable, 'map_draw.py', 'maps/Map.xlsx', '--ids', rooms_str, '--out-sheet', sheet_name],
            capture_output=True,
            text=True,
        )

        if result.returncode != 0:
            print(f'  ERROR: {result.stderr}')

    # Add abilities text to each sphere sheet
    print()
    print('Adding abilities to sheet headers...')

    from openpyxl import load_workbook

    wb = load_workbook('maps/Map.xlsx')

    for sphere_data in spheres:
        sphere_num = sphere_data['sphere']
        abilities = sphere_data['abilities']
        sheet_name = f"Sphere {sphere_num}"

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws['A1'] = f"Abilities: {abilities}"
            print(f'  {sheet_name}: "{abilities}"')

    wb.save('maps/Map.xlsx')
    print()
    print('Done!')


if __name__ == '__main__':
    main()
